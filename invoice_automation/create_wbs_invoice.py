r"""
Fill SAP Supplier Invoice template from North Sky Vendor Invoice (Master).

Mapping requested:
  Source Invoice Number  -> Template Reference
  Source Invoice Amount  -> Template Gross Invoice Amount in Document Currency
  Source Invoice Date    -> Template Document Date / Invoice Date
  Template Document Type -> default "NT"
  Transaction -> default 1
  Header Company Code and G/L item Company Code -> default 4600
  G/L Account / Account -> default 51000100
  Debit/Credit -> default S
  Tax Code -> default I0
  G/L item Amount in Document Currency -> source Invoice Amount
  G/L item Item Text -> default Supplies
  Source Vendor "vendor_name - vendor_id" -> Template Invoicing Party (vendor_id only)
  Source Contract ID + WBS Area code + WBS Supervisor code + vendor_id -> Template WBS Element

Filter rule:
  Only export source rows where Status = "Need Non-PO Invoice Created" and Contract ID is not blank and contains at least 5 digits.

Default behavior on your computer:
  - Master input:
      C:\PYTHON\PDFInvoiceExtract\invoice_automation\clients\northsky_comm\MasterExcel\North Sky Vendor Invoice (Master).xlsx
  - Template input:
      C:\PYTHON\PDFInvoiceExtract\invoice_automation\clients\northsky_comm\templates\Invoice_Template_no_PO.XLSX
  - Output Excel goes to the same MasterExcel directory and overwrites:
      invoice_template_filled.xlsx

You can still pass paths manually:
  python fill_invoice_template_from_master.py master.xlsx template.xlsx output.xlsx

Requires:
  pip install openpyxl
"""

from __future__ import annotations

import argparse
import re
from copy import copy
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, Iterable, Optional

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


DEFAULT_DOCUMENT_TYPE = "NT"
DEFAULT_MASTER_FILE = r"C:\PYTHON\PDFInvoiceExtract\invoice_automation\clients\northsky_comm\MasterExcel\North Sky Vendor Invoice (Master).xlsx"
DEFAULT_TEMPLATE_NAME = "Invoice_Template_no_PO.XLSX"
DEFAULT_OUTPUT_NAME = "invoice_template_filled.xlsx"
DEFAULT_TRANSACTION_TYPE = 1
DEFAULT_COMPANY_CODE = "4600"
DEFAULT_GL_ACCOUNT = "51000100"
DEFAULT_DEBIT_CREDIT = "S"
DEFAULT_TAX_CODE = "I0"
DEFAULT_ITEM_TEXT = "Supplies"


def clean_header(value: Any) -> str:
    """Normalize Excel header text for matching."""
    if value is None:
        return ""
    return str(value).replace("\r", " ").replace("\n", " ").strip().lower()


def find_header_row(ws, required_headers: Iterable[str], max_rows: int = 20) -> int:
    """Find the row containing all required header names."""
    required = {h.lower() for h in required_headers}
    for row in range(1, min(ws.max_row, max_rows) + 1):
        values = {clean_header(ws.cell(row=row, column=col).value) for col in range(1, ws.max_column + 1)}
        if required.issubset(values):
            return row
    raise ValueError(f"Could not find header row with headers: {sorted(required)}")


def header_map(ws, header_row: int) -> Dict[str, int]:
    """Return normalized header -> column number."""
    result: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = clean_header(ws.cell(row=header_row, column=col).value)
        if key and key not in result:
            result[key] = col
    return result


def find_template_row(ws) -> int:
    """
    SAP template has technical headers like SUPPLIERINVOICEIDBYINVCGPARTY,
    DOCUMENTDATE, ACCOUNTINGDOCUMENTTYPE, INVOICEGROSSAMOUNT.
    """
    required = {
        "supplierinvoiceidbyinvcgparty",
        "documentdate",
        "accountingdocumenttype",
        "invoicegrossamount",
    }
    for row in range(1, min(ws.max_row, 20) + 1):
        values = {clean_header(ws.cell(row=row, column=col).value) for col in range(1, ws.max_column + 1)}
        if required.issubset(values):
            return row
    raise ValueError("Could not find SAP template technical header row.")


def get_first_existing(mapping: Dict[str, int], *names: str) -> Optional[int]:
    """Find the first matching column by normalized header name."""
    for name in names:
        col = mapping.get(clean_header(name))
        if col:
            return col
    return None


def get_all_existing(ws, header_row: int, *names: str) -> list[int]:
    """Find all matching columns by normalized header name, including duplicate template columns."""
    targets = {clean_header(name) for name in names}
    cols: list[int] = []
    for col in range(1, ws.max_column + 1):
        if clean_header(ws.cell(row=header_row, column=col).value) in targets:
            cols.append(col)
    return cols


def set_many(ws, row: int, cols: Iterable[int], value: Any) -> None:
    """Write the same value to multiple target columns."""
    for col in cols:
        ws.cell(row=row, column=col).value = value


def to_excel_date(value: Any) -> Any:
    """Convert source date to a real date where possible."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return value
    text = str(value).strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y", "%m-%d-%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return value


def clean_text(value: Any) -> str:
    """Return a trimmed string, converting blank/None to empty string."""
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"none", "nan"}:
        return ""
    return text


def normalize_contract_id(value: Any) -> str:
    """Keep Contract ID as text, but avoid 21310.0 when Excel stores it as a number."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean_text(value)


def normalize_status(value: Any) -> str:
    """Normalize Status for forgiving comparisons."""
    text = clean_text(value).lower()
    text = text.replace("’", "'")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def is_need_non_po_invoice_created_status(value: Any) -> bool:
    """Return True for Status values like 'Need Non-PO Invoice Created'."""
    return normalize_status(value) in {
        "need non po invoice created",
        "need nonpo invoice created",
    }


def contract_id_has_number(value: Any) -> bool:
    """Contract ID must not be blank and must contain at least 5 digits."""
    contract_id = normalize_contract_id(value)
    digit_count = len(re.findall(r"\d", contract_id))
    return bool(contract_id and digit_count >= 5)


def normalize_identifier(value: Any) -> str:
    """Keep invoice/reference IDs as text and avoid Excel-looking values like 18920.0."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = clean_text(value)
    if text.endswith(".0"):
        maybe_number = text[:-2]
        if maybe_number.replace(",", "").isdigit():
            return maybe_number
    return text


def extract_vendor_id(vendor_value: Any) -> str:
    """
    Extract vendor id from source Vendor text like:
      ADVANCED UNDERGROUND UTILITY LOCATING INC-V02075
      KINNAN ENGINEERING - S00262
    Returns only V02075 / S00262.
    """
    text = clean_text(vendor_value)
    if not text:
        return ""

    # Prefer the token after the last dash. This handles vendor names that contain spaces.
    if "-" in text:
        candidate = text.rsplit("-", 1)[-1].strip()
    else:
        candidate = text.split()[-1].strip()

    return candidate


def first_code(value: Any) -> str:
    """
    Return only the short code before the first dash.

    Examples:
      "WA0 - Washington" -> "WA0"
      "829151-Robert Michael Thoman" -> "829151"
    """
    text = clean_text(value)
    if not text:
        return ""
    return text.split("-", 1)[0].strip()


def build_wbs_element(contract_id: Any, wbs_area: Any, wbs_supervisor: Any, vendor_id: Any) -> str:
    """
    Build WBS Element using only short WBS codes, not descriptions/names.

    Keeps:
      Contract ID as-is, e.g. 21310
      WBS Area code only, e.g. WA0 from "WA0 - Washington"
      WBS Supervisor code only, e.g. 829151 from "829151-Robert Michael Thoman"
      Vendor ID as-is, e.g. V02075
    """
    parts = [
        normalize_contract_id(contract_id),
        first_code(wbs_area),
        first_code(wbs_supervisor),
        clean_text(vendor_id),
    ]
    return "-".join(part for part in parts if part)


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    """Copy template row style/formulas/data validation-friendly formats to a new output row."""
    if target_row == source_row:
        return
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=source_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)
        if src.border:
            dst.border = copy(src.border)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.font:
            dst.font = copy(src.font)


def fill_template(master_file: Path, template_file: Path, output_file: Path) -> None:
    master_wb = load_workbook(master_file, data_only=True)
    master_ws = master_wb.active

    source_header_row = find_header_row(
        master_ws,
        required_headers=["Invoice Number", "Invoice Date", "Invoice Amount", "Vendor", "Contract ID", "Status"],
    )
    source_cols = header_map(master_ws, source_header_row)

    invoice_no_col = get_first_existing(source_cols, "Invoice Number")
    invoice_date_col = get_first_existing(source_cols, "Invoice Date")
    invoice_amount_col = get_first_existing(source_cols, "Invoice Amount")
    workflow_col = get_first_existing(source_cols, "Workflow Category ID", "Workflow Category")
    status_col = get_first_existing(source_cols, "Status")
    vendor_col = get_first_existing(source_cols, "Vendor")
    contract_id_col = get_first_existing(source_cols, "Contract ID")
    wbs_area_col = get_first_existing(source_cols, "WBS (Area)", "WBS Area")
    wbs_supervisor_col = get_first_existing(source_cols, "WBS (Supervisor)", "WBS Supervisor")

    if not invoice_no_col or not invoice_date_col or not invoice_amount_col:
        raise ValueError("Source file is missing Invoice Number, Invoice Date, or Invoice Amount column.")
    if not status_col:
        raise ValueError("Source file is missing Status column.")

    template_wb = load_workbook(template_file)
    template_ws = template_wb.active

    technical_header_row = find_template_row(template_ws)
    template_cols = header_map(template_ws, technical_header_row)
    data_start_row = technical_header_row + 2  # row after user-friendly label row

    reference_col = get_first_existing(template_cols, "SUPPLIERINVOICEIDBYINVCGPARTY")
    gross_amount_col = get_first_existing(template_cols, "INVOICEGROSSAMOUNT")
    document_date_col = get_first_existing(template_cols, "DOCUMENTDATE")
    document_type_col = get_first_existing(template_cols, "ACCOUNTINGDOCUMENTTYPE")
    invoicing_party_col = get_first_existing(template_cols, "INVOICINGPARTY")
    wbs_element_col = get_first_existing(template_cols, "WBSELEMENT")

    if not all([reference_col, gross_amount_col, document_date_col, document_type_col]):
        raise ValueError("Template is missing one or more target columns.")

    # Optional required SAP columns often needed in this template.
    label_cols = header_map(template_ws, technical_header_row + 1)
    invoice_id_col = get_first_existing(template_cols, "3") or get_first_existing(label_cols, "*Invoice ID", "Invoice ID")
    company_code_cols = get_all_existing(template_ws, technical_header_row, "COMPANYCODE")
    transaction_col = get_first_existing(template_cols, "SUPPLIERINVOICETRANSACTIONTYPE")
    currency_col = get_first_existing(template_cols, "DOCUMENTCURRENCY")

    # G/L Account item columns
    gl_account_col = get_first_existing(template_cols, "GLACCOUNT")
    line_item_text_col = get_first_existing(template_cols, "SUPPLIERINVOICEITEMTEXT")
    debit_credit_col = get_first_existing(template_cols, "DEBITCREDITCODE")
    line_amount_col = get_first_existing(template_cols, "SUPPLIERINVOICEITEMAMOUNT")
    tax_code_col = get_first_existing(template_cols, "TAXCODE")

    output_row = data_start_row
    invoice_id = 1

    for source_row in range(source_header_row + 1, master_ws.max_row + 1):
        invoice_no = master_ws.cell(source_row, invoice_no_col).value
        invoice_amount = master_ws.cell(source_row, invoice_amount_col).value
        invoice_date = master_ws.cell(source_row, invoice_date_col).value
        status_value = master_ws.cell(source_row, status_col).value if status_col else None
        vendor_value = master_ws.cell(source_row, vendor_col).value if vendor_col else None
        contract_id = master_ws.cell(source_row, contract_id_col).value if contract_id_col else None

        # Skip blank source rows.
        if invoice_no in (None, "") and invoice_amount in (None, "") and invoice_date in (None, ""):
            continue

        # Requested filter: only export rows with Status = Need Non-PO Invoice Created
        # and Contract ID not blank / containing at least 5 digits.
        if not is_need_non_po_invoice_created_status(status_value):
            continue
        if not contract_id_has_number(contract_id):
            continue

        wbs_area = master_ws.cell(source_row, wbs_area_col).value if wbs_area_col else None
        wbs_supervisor = master_ws.cell(source_row, wbs_supervisor_col).value if wbs_supervisor_col else None
        vendor_id = extract_vendor_id(vendor_value)
        wbs_element = build_wbs_element(contract_id, wbs_area, wbs_supervisor, vendor_id)

        copy_row_style(template_ws, data_start_row, output_row)

        # Requested mappings
        template_ws.cell(output_row, reference_col).value = normalize_identifier(invoice_no) or None
        template_ws.cell(output_row, gross_amount_col).value = invoice_amount
        template_ws.cell(output_row, document_date_col).value = to_excel_date(invoice_date)
        template_ws.cell(output_row, document_type_col).value = DEFAULT_DOCUMENT_TYPE
        if invoicing_party_col:
            template_ws.cell(output_row, invoicing_party_col).value = vendor_id or None
        if wbs_element_col:
            template_ws.cell(output_row, wbs_element_col).value = wbs_element or None

        # Helpful defaults for SAP template. Delete these lines if you do not want defaults.
        if invoice_id_col:
            template_ws.cell(output_row, invoice_id_col).value = invoice_id
        if company_code_cols:
            set_many(template_ws, output_row, company_code_cols, DEFAULT_COMPANY_CODE)
        if transaction_col:
            template_ws.cell(output_row, transaction_col).value = DEFAULT_TRANSACTION_TYPE
        if currency_col:
            template_ws.cell(output_row, currency_col).value = "USD"

        # Requested G/L Account item defaults/mappings.
        if gl_account_col:
            template_ws.cell(output_row, gl_account_col).value = DEFAULT_GL_ACCOUNT
        if debit_credit_col:
            template_ws.cell(output_row, debit_credit_col).value = DEFAULT_DEBIT_CREDIT
        if tax_code_col:
            template_ws.cell(output_row, tax_code_col).value = DEFAULT_TAX_CODE
        if line_amount_col:
            template_ws.cell(output_row, line_amount_col).value = invoice_amount
        if line_item_text_col:
            template_ws.cell(output_row, line_item_text_col).value = DEFAULT_ITEM_TEXT

        # Keep dates and amount formatted correctly.
        template_ws.cell(output_row, document_date_col).number_format = "mm/dd/yy"
        template_ws.cell(output_row, gross_amount_col).number_format = '#,##0.00'
        if line_amount_col:
            template_ws.cell(output_row, line_amount_col).number_format = '#,##0.00'
        if invoicing_party_col:
            template_ws.cell(output_row, invoicing_party_col).number_format = '@'
        if wbs_element_col:
            template_ws.cell(output_row, wbs_element_col).number_format = '@'
        if gl_account_col:
            template_ws.cell(output_row, gl_account_col).number_format = '@'
        if line_item_text_col:
            template_ws.cell(output_row, line_item_text_col).number_format = '@'

        invoice_id += 1
        output_row += 1

    template_wb.save(output_file)
    print(f"Created: {output_file}")
    print(f"Rows written: {invoice_id - 1}")


def default_template_for_master(master_file: Path) -> Path:
    """Default template location: sibling templates folder under the client root."""
    # master_file is usually:
    #   ...\clients\northsky_comm\MasterExcel\North Sky Vendor Invoice (Master).xlsx
    # Template is usually:
    #   ...\clients\northsky_comm\templates\Invoice_Template_no_PO.XLSX
    client_root = master_file.parent.parent
    return client_root / "templates" / DEFAULT_TEMPLATE_NAME


def default_output_for_master(master_file: Path) -> Path:
    """Write output Excel into the same directory as the Master Excel file."""
    return master_file.parent / DEFAULT_OUTPUT_NAME


def main() -> None:
    parser = argparse.ArgumentParser(description="Fill SAP invoice template from North Sky Vendor Invoice Master Excel.")
    parser.add_argument("master_file", nargs="?", default=DEFAULT_MASTER_FILE)
    parser.add_argument("template_file", nargs="?", default=None)
    parser.add_argument("output_file", nargs="?", default=None)
    args = parser.parse_args()

    master_file = Path(args.master_file)
    template_file = Path(args.template_file) if args.template_file else default_template_for_master(master_file)
    output_file = Path(args.output_file) if args.output_file else default_output_for_master(master_file)

    output_file.parent.mkdir(parents=True, exist_ok=True)

    print(f"Master input:  {master_file}")
    print(f"Template input:{template_file}")
    print(f"Output Excel:  {output_file}")

    fill_template(master_file, template_file, output_file)


if __name__ == "__main__":
    main()
