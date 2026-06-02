import tempfile
import unittest
from datetime import datetime
from pathlib import Path

import openpyxl

from invoice_automation.app.excel_writer import (
    set_if_exists,
    write_invoice_row,
    write_invoices_to_vendor_template_batches,
)


class ExcelWriterTests(unittest.TestCase):
    def test_set_if_exists_defaults_missing_value_to_blank(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = {"OPTIONALFIELD": 1}

        set_if_exists(ws, 2, headers, "OPTIONALFIELD")

        self.assertEqual(ws.cell(row=2, column=1).value, "")

    def test_write_invoice_row_writes_transaction_type(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        header_names = [
            "SUPPLIERINVOICEITEM",
            "INVOICEID",
            "ITEMID",
            "COMPANYCODE",
            "SUPPLIERINVOICETRANSACTIONTYPE",
            "INVOICINGPARTY",
            "SUPPLIERINVOICEIDBYINVCGPARTY",
            "DOCUMENTDATE",
            "POSTINGDATE",
            "ACCOUNTINGDOCUMENTTYPE",
            "ACCOUNTINGDOCUMENTHEADERTEXT",
            "DOCUMENTCURRENCY",
            "INVOICEGROSSAMOUNT",
            "GLACCOUNT",
            "DEBITCREDITCODE",
            "SUPPLIERINVOICEITEMAMOUNT",
            "TAXCODE",
            "SUPPLIERINVOICEITEMTEXT",
            "TAXJURISDICTION",
            "COSTCENTER",
            "DOCUMENTITEMTEXT",
        ]
        for col, name in enumerate(header_names, start=1):
            ws.cell(row=1, column=col).value = name
        headers = {name: col for col, name in enumerate(header_names, start=1)}

        write_invoice_row(
            ws=ws,
            row=2,
            headers=headers,
            invoice={
                "amount": "123.45",
                "invoice_date": "2026-05-31",
                "invoice_number": "INV-1",
                "vendor_name": "Vendor",
                "TaxCenterID": "TX001",
            },
            vendor_config={
                "company_code": "1000",
                "vendor_code": "V100",
                "gl_account": "5000",
                "tax_code": "I0",
                "item_text": "Repairs",
                "cost_center": "CC100",
            },
            excel_config={
                "supplier_invoice_transaction_type": "2",
                "accounting_document_type": "NS",
                "document_currency": "USD",
            },
            line_number=1,
        )

        self.assertEqual(
            ws.cell(row=2, column=headers["SUPPLIERINVOICETRANSACTIONTYPE"]).value,
            "2",
        )
        self.assertEqual(
            ws.cell(row=2, column=headers["SUPPLIERINVOICEIDBYINVCGPARTY"]).value,
            "INV-1",
        )
        self.assertEqual(
            ws.cell(row=2, column=headers["INVOICEGROSSAMOUNT"]).value,
            "123.45",
        )
        self.assertIsInstance(
            ws.cell(row=2, column=headers["DOCUMENTDATE"]).value,
            datetime,
        )

    def test_append_output_mode_adds_records_to_existing_workbook(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            template_file = root / "template.xlsx"
            output_file = root / "G_drive" / "invoice_output.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"
            headers = [
                "SUPPLIERINVOICEITEM",
                "COMPANYCODE",
                "SUPPLIERINVOICEIDBYINVCGPARTY",
                "DOCUMENTDATE",
                "INVOICEGROSSAMOUNT",
            ]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col).value = header
            wb.save(template_file)

            vendor_config = {
                "template_file": str(template_file),
                "output_folder": str(output_file.parent),
                "output_file": str(output_file),
                "company_code": "1000",
            }
            excel_config = {
                "sheet_name": "Data",
                "output_mode": "append",
            }

            write_invoices_to_vendor_template_batches(
                invoices=[
                    {
                        "amount": "100.00",
                        "invoice_date": "2026-05-31",
                        "invoice_number": "INV-1",
                    }
                ],
                vendor_config=vendor_config,
                excel_config=excel_config,
                client_root=root,
            )
            write_invoices_to_vendor_template_batches(
                invoices=[
                    {
                        "amount": "200.00",
                        "invoice_date": "2026-06-01",
                        "invoice_number": "INV-2",
                    }
                ],
                vendor_config=vendor_config,
                excel_config=excel_config,
                client_root=root,
            )

            output_wb = openpyxl.load_workbook(output_file)
            output_ws = output_wb["Data"]

            self.assertEqual(output_ws.cell(row=2, column=1).value, 1)
            self.assertEqual(output_ws.cell(row=3, column=1).value, 2)
            self.assertEqual(output_ws.cell(row=2, column=3).value, "INV-1")
            self.assertEqual(output_ws.cell(row=3, column=3).value, "INV-2")
            self.assertEqual(output_ws.cell(row=3, column=5).value, "200.00")


if __name__ == "__main__":
    unittest.main()
