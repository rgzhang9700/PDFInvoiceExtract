from pathlib import Path
from datetime import datetime
import yaml

from app.email_downloader import download_invoice_pdfs_for_accounts
from app.pdf_text import extract_pdf_text
from app.file_mover import move_failed_pdf
from app.logger import ProcessingLogger

from app.parsers.valvoline_parser import ValvolineParser
from app.parsers.fleetpride_parser import FleetPrideParser

from app.excel_writer import write_invoices_to_vendor_template_batches


PARSER_CLASSES = {
    "ValvolineParser": ValvolineParser,
    "FleetPrideParser": FleetPrideParser,
}


def load_parser_rules(client_root):
    """
    Read parser rules from:
        clients/northsky_comm/templates/SupplierLists.xlsx
        sheet: Parsers

    Expected columns:
        Vendor Name | Supplier | Parser

    If no match is found later, default parser is ValvolineParser.
    """
    supplier_file = client_root / "templates" / "SupplierLists.xlsx"

    if not supplier_file.exists():
        print(f"Supplier parser list not found, using default ValvolineParser: {supplier_file}")
        return []

    import openpyxl

    wb = openpyxl.load_workbook(supplier_file, data_only=True)

    if "Parsers" not in wb.sheetnames:
        print(f"Sheet 'Parsers' not found in {supplier_file}, using default ValvolineParser")
        return []

    ws = wb["Parsers"]

    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            headers[str(value).strip().upper()] = col

    vendor_col = headers.get("VENDOR NAME")
    parser_col = headers.get("PARSER")
    supplier_col = headers.get("SUPPLIER")

    if not vendor_col or not parser_col:
        print("SupplierLists.xlsx Parsers sheet must have 'Vendor Name' and 'Parser' columns.")
        return []

    rules = []

    for row in range(2, ws.max_row + 1):
        vendor_name = ws.cell(row=row, column=vendor_col).value
        parser_name = ws.cell(row=row, column=parser_col).value
        supplier = ws.cell(row=row, column=supplier_col).value if supplier_col else ""

        vendor_name = str(vendor_name or "").strip()
        parser_name = str(parser_name or "").strip()
        supplier = str(supplier or "").strip()

        if vendor_name and parser_name:
            rules.append({
                "vendor_name": vendor_name.upper(),
                "parser_name": parser_name,
                "supplier": supplier,
            })

    print(f"Loaded {len(rules)} parser rules from {supplier_file}")
    return rules


def detect_parser(text, parser_rules=None):
    """
    Pick parser using vendor name.

    Simple logic:
        1. Use default ValvolineParser to find vendor name.
        2. Match that vendor name against SupplierLists.xlsx / Parsers sheet.
        3. Use the Parser from the matching row.
        4. If nothing matches, default to ValvolineParser.
    """
    text = text or ""
    parser_rules = parser_rules or []

    default_parser = ValvolineParser()

    vendor_name = ""
    if hasattr(default_parser, "_find_vendor_name"):
        vendor_name = default_parser._find_vendor_name(text)

    # Extra fallback if _find_vendor_name() did not include FleetPride or other names.
    if not vendor_name:
        upper_text = text.upper()
        for rule in parser_rules:
            rule_vendor = rule.get("vendor_name", "")
            if rule_vendor and rule_vendor.upper() in upper_text:
                vendor_name = rule_vendor
                break

    vendor_name_upper = (vendor_name or "").upper()
    print(f"Detected vendor name: {vendor_name}")

    for rule in parser_rules:
        rule_vendor = rule.get("vendor_name", "").upper()
        parser_name = rule.get("parser_name", "")

        if rule_vendor and vendor_name_upper and rule_vendor in vendor_name_upper:
            parser_class = PARSER_CLASSES.get(parser_name, ValvolineParser)
            print(f"Matched parser from SupplierLists: {rule_vendor} -> {parser_class.__name__}")
            return parser_class()

    print("No parser match found. Using default ValvolineParser.")
    return ValvolineParser()


def run_client(config_path: Path):
    config_path = Path(config_path)
    client_root = config_path.parent

    with open(config_path, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)

    parser_rules = load_parser_rules(client_root)

    # Keep ProcessingLogger only for run_summary.
    # Do not call write_error() or write_success(), so no separate error_log/summary_log is created.
    logger = ProcessingLogger(
        resolve_path(client_root, config["logging"]["log_folder"]),
        config["logging"],
    )

    if config.get("email", {}).get("enabled", False):
        download_invoice_pdfs_for_accounts(
            email_config=config["email"],
            client_root=client_root,
        )

    grand_total_found = 0
    grand_success = 0
    grand_failed = 0

    # One Excel file per run: collect all invoices first, then write once.
    all_invoices = []
    output_vendor_config = None
    vendor_summaries = {}

    vendor_base_config = {
        key: value
        for key, value in config["vendors"].items()
        if not isinstance(value, dict)
    }

    # If several vendor sections use the same input folder, do not process
    # the same files again and again.
    processed_input_folders = set()

    for vendor_folder, vendor_specific_config in config["vendors"].items():
        # Skip shared vendor settings like input_folder/template_file/output_folder.
        if not isinstance(vendor_specific_config, dict):
            continue

        # Merge shared config + vendor-specific config.
        # IMPORTANT:
        # Do NOT add vendor_folder to input_folder.
        # Your files are directly in ./downloads, not ./downloads/<vendor_folder>.
        vendor_config = dict(vendor_base_config)
        vendor_config.update(vendor_specific_config)

        input_folder = resolve_path(client_root, vendor_config["input_folder"]).resolve()

        if input_folder in processed_input_folders:
            print(f"Skipping duplicate input folder for {vendor_folder}: {input_folder}")
            continue

        processed_input_folders.add(input_folder)

        invoices, successful_records, failed_records, summary = parse_vendor_folder(
            vendor_folder=vendor_folder,
            vendor_config=vendor_config,
            processing_config=config["processing"],
            client_root=client_root,
            logger=logger,
            parser_rules=parser_rules,
        )

        vendor_summaries[vendor_folder] = summary
        grand_total_found += summary["total_files_found"]
        grand_failed += summary["failed_count"]

        if invoices:
            all_invoices.extend(invoices)

            # All vendors use the same template. Use the first vendor config
            # as the output/template/GL config for this run.
            if output_vendor_config is None:
                output_vendor_config = vendor_config

    if all_invoices:
        try:
            write_invoices_to_vendor_template_batches(
                invoices=all_invoices,
                vendor_config=output_vendor_config,
                excel_config=config["excel"],
                client_root=client_root,
            )

            # Move PDFs only after the single Excel write succeeds.
            moved_by_vendor = move_successful_invoices_after_excel(
                invoices=all_invoices,
                vendor_folder="ALL_VENDORS",
                processing_config=config["processing"],
                client_root=client_root,
                logger=logger,
            )

            for vendor_folder, moved_count in moved_by_vendor.items():
                if vendor_folder in vendor_summaries:
                    vendor_summaries[vendor_folder]["success_count"] += moved_count
                grand_success += moved_count

        except Exception as e:
            print(f"EXCEL WRITE FAILED for one-file run: {e}")

            # Do not move PDFs to processed if Excel failed.
            # Leave them in input folder so they can be fixed/reprocessed.
            for invoice in all_invoices:
                print_file_result({
                    "pdf_file": invoice.get("pdf_file", ""),
                    "vendor_name": invoice.get("vendor_name", ""),
                    "invoice_number": invoice.get("invoice_number", ""),
                    "invoice_date": invoice.get("invoice_date", ""),
                    "amount": invoice.get("amount", ""),
                    "status": "excel_failed_not_moved",
                    "error": str(e),
                })

    for vendor_folder, summary in vendor_summaries.items():
        logger.write_run_summary({
            "run_time": datetime.now().isoformat(timespec="seconds"),
            "vendor_folder": vendor_folder,
            "total_files_found": summary["total_files_found"],
            "success_count": summary["success_count"],
            "failed_count": summary["failed_count"],
            "total_invoice_files_processed": (
                summary["success_count"]
                + summary["failed_count"]
            ),
        })

    print("===================================")
    print(f"CLIENT: {config.get('client', {}).get('name', 'Unknown')}")
    print("RUN TOTAL")
    print(f"Total files found: {grand_total_found}")
    print(f"Success: {grand_success}")
    print(f"Failed: {grand_failed}")
    print(f"Total invoice files processed: {grand_success + grand_failed}")
    print("===================================")


def parse_vendor_folder(
    vendor_folder,
    vendor_config,
    processing_config,
    client_root,
    logger,
    parser_rules=None,
):
    input_folder = resolve_path(client_root, vendor_config["input_folder"])

    # Only look in the current download folder.
    # Do NOT scan PROCESSED or any other subdirectory.
    pdf_files = sorted(
        f for f in input_folder.iterdir()
        if f.is_file() and f.suffix.lower() == ".pdf"
    )

    summary = {
        "total_files_found": len(pdf_files),
        "success_count": 0,
        "failed_count": 0,
        "duplicate_count": 0,
    }

    invoices = []
    successful_records = []
    failed_records = []

    if not pdf_files:
        return invoices, successful_records, failed_records, summary

    for pdf_file in pdf_files:
        print(f"Processing {vendor_folder}: {pdf_file.name}")

        try:
            text, pdf_type = extract_pdf_text(pdf_file)

            parser = detect_parser(text, parser_rules=parser_rules)
            print(f"Using parser: {parser.__class__.__name__}")

            invoice = parser.parse(text)

            invoice["vendor_folder"] = vendor_folder
            invoice["pdf_type"] = pdf_type
            invoice["pdf_file"] = pdf_file.name
            invoice["source_pdf_path"] = str(pdf_file)
            invoice["status"] = "parsed_waiting_for_excel"
            invoice["error"] = ""

            validate_invoice(invoice)

            successful_records.append(invoice.copy())

            # Do not move to processed here.
            # The PDF stays in input until Excel write succeeds.
            invoices.append(invoice)

            print_file_result(invoice)

        except Exception as e:
            summary["failed_count"] += 1

            moved_path = ""
            if processing_config.get("move_failed", True):
                moved_path = move_failed_pdf(
                    pdf_file=pdf_file,
                    vendor_folder=vendor_folder,
                    failed_root=resolve_path(
                        client_root,
                        processing_config.get("failed_folder", "./error"),
                    ),
                    append_timestamp_if_exists=processing_config.get(
                        "append_timestamp_if_exists",
                        True,
                    ),
                )

            failed_record = {
                "vendor_folder": vendor_folder,
                "pdf_file": pdf_file.name,
                "processed_file_path": moved_path,
                "status": "failed",
                "error": str(e),
            }
            failed_records.append(failed_record)

            # Do not write separate error_log; failures are counted in run_summary only.

            print_file_result({
                "pdf_file": pdf_file.name,
                "vendor_name": "",
                "invoice_number": "",
                "invoice_date": "",
                "amount": "",
                "status": "failed",
                "error": str(e),
            })

            print(f"FAILED {pdf_file.name}: {e}")

    return invoices, successful_records, failed_records, summary


def move_successful_invoices_after_excel(
    invoices,
    vendor_folder,
    processing_config,
    client_root,
    logger,
):
    moved_by_vendor = {}

    for invoice in invoices:
        source_file = Path(invoice["source_pdf_path"])
        invoice_vendor_folder = invoice.get("vendor_folder", vendor_folder)

        # NEW:
        # Move passed/processed source file into PROCESSED folder
        # inside the same current download folder.
        processed_folder = source_file.parent / "PROCESSED"
        processed_folder.mkdir(parents=True, exist_ok=True)

        target_file = processed_folder / source_file.name

        # If same filename exists, append timestamp.
        if target_file.exists():
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            target_file = processed_folder / f"{source_file.stem}_{timestamp}{source_file.suffix}"

        source_file.rename(target_file)

        invoice["processed_file_path"] = str(target_file)
        invoice["status"] = "success"
        invoice["error"] = ""

        # Do not write separate success/summary logs; run_summary is written in run_client().

        moved_by_vendor[invoice_vendor_folder] = moved_by_vendor.get(invoice_vendor_folder, 0) + 1

    return moved_by_vendor


def print_file_result(invoice):
    postcode_lookup = (
        invoice.get("postcode_lookup")
        or invoice.get("tax_center_id")
        or invoice.get("TaxCenterID")
        or invoice.get("taxcenter_id")
        or ""
    )

    print(
        f"File: {invoice.get('pdf_file', '')} | "
        f"Vendor: {invoice.get('vendor_name', '')} | "
        f"Invoice No: {invoice.get('invoice_number', '')} | "
        f"Invoice Date: {invoice.get('invoice_date', '')} | "
        f"Total Amount: {invoice.get('amount', '')} | "
        f"postcode_lookup: {postcode_lookup}"
    )


def validate_invoice(invoice):
    required_fields = [
        "invoice_number",
        "invoice_date",
        "amount",
    ]

    missing = [field for field in required_fields if not invoice.get(field)]

    if missing:
        raise ValueError(f"Missing required fields: {missing}")

    try:
        amount = float(invoice.get("amount", 0))
        if amount <= 0:
            raise ValueError(f"Invalid invoice amount: {amount}")
    except Exception:
        raise ValueError(f"Invalid invoice amount: {invoice.get('amount')}")


def resolve_path(client_root, path_value):
    p = Path(path_value)

    if p.is_absolute() or ":" in str(path_value):
        return p

    return client_root / p
