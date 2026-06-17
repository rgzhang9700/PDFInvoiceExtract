from pathlib import Path
import shutil
import math
import re
from datetime import datetime
import openpyxl


def write_invoices_to_vendor_template_batches(invoices, vendor_config, excel_config, client_root):
    template_file = resolve_path(client_root, vendor_config["template_file"])

    if not template_file.exists():
        raise FileNotFoundError(f"Vendor Excel template not found: {template_file}")

    if not invoices:
        print("No invoices to write.")
        return

    max_records = int(excel_config.get("max_records_per_file", 50))
    file_prefix = vendor_config.get("file_prefix", "invoice_load")

    # NEW:
    # Write the output Excel into the same folder as the source PDF/image files.
    # If invoices come from multiple folders, create one output Excel set per folder.
    invoices_by_folder = {}

    for invoice in invoices:
        source_path = invoice.get("source_pdf_path") or invoice.get("source_image_path") or invoice.get("pdf_file", "")
        source_folder = Path(source_path).parent if source_path else resolve_path(client_root, vendor_config["output_folder"])
        output_folder = source_folder / "PROCESSED"
        invoices_by_folder.setdefault(output_folder, []).append(invoice)

    for output_folder, folder_invoices in invoices_by_folder.items():
        output_folder = Path(output_folder)
        output_folder.mkdir(parents=True, exist_ok=True)

        total_batches = math.ceil(len(folder_invoices) / max_records)

        for batch_index in range(total_batches):
            batch_records = folder_invoices[batch_index * max_records:(batch_index + 1) * max_records]
            output_file = output_folder / f"{file_prefix}_{batch_index + 1:03d}.xlsx"

            shutil.copy(template_file, output_file)

            wb = openpyxl.load_workbook(output_file)
            sheet_name = excel_config.get("sheet_name", "Data")

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in vendor template: {template_file}")

            ws = wb[sheet_name]
            header_row = find_header_row(ws)
            headers = build_header_map(ws, header_row)
            start_row = find_next_empty_row(ws, header_row + 1)

            for index, invoice in enumerate(batch_records):
                row = start_row + index

                # Invoice/row ID increases 1, 2, 3 within each Excel output file.
                line_number = index + 1

                write_invoice_row(
                    ws=ws,
                    row=row,
                    headers=headers,
                    invoice=invoice,
                    vendor_config=vendor_config,
                    excel_config=excel_config,
                    line_number=line_number,
                )

            wb.save(output_file)
            print(f"Excel created: {output_file} with {len(batch_records)} records")


def write_processing_summary_excel(successful_records, failed_records, output_file):
    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    success_ws = wb.active
    success_ws.title = "Successful Parses"
    success_headers = [
        "filename",
        "invoice_number",
        "invoice_date",
        "total_amount",
        "post_code",
        "vendor_folder",
        "vendor_name",
        "status",
    ]
    success_ws.append(success_headers)

    for record in successful_records:
        success_ws.append([
            record.get("pdf_file", ""),
            record.get("invoice_number", ""),
            record.get("invoice_date", ""),
            record.get("amount", ""),
            get_invoice_post_code(record),
            record.get("vendor_folder", ""),
            record.get("vendor_name", ""),
            record.get("status", ""),
        ])

    failed_ws = wb.create_sheet("Failed Parses")
    failed_headers = ["filename", "vendor_folder", "status", "error", "processed_file_path"]
    failed_ws.append(failed_headers)

    for record in failed_records:
        failed_ws.append([
            record.get("pdf_file", ""),
            record.get("vendor_folder", ""),
            record.get("status", "failed"),
            record.get("error", ""),
            record.get("processed_file_path", ""),
        ])

    for ws in (success_ws, failed_ws):
        for column_cells in ws.columns:
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            ws.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 60)

    wb.save(output_file)
    print(f"Processing summary Excel created: {output_file}")


def get_invoice_post_code(invoice):
    for key in (
        "ship_to_postcode",
        "service_center_postcode",
        "vendor_postcode",
        "post_code",
        "postcode",
    ):
        value = invoice.get(key, "")
        if value:
            return value
    return ""




def to_number(value):
    """Convert invoice amount text to a real Excel number.

    Examples supported:
    49.10, "$49.10", "1,148.98", "s1,148. 98", "Total $2,185.00".
    Returns None when no usable number is found, so Excel cell stays blank.
    """
    if value is None or value == "":
        return None

    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    if not s:
        return None

    # Common OCR fixes: s/S before amount can mean $, and spaces may appear around decimal point.
    s = s.replace("$", "").replace(",", "")
    s = re.sub(r"(?i)\bs(?=\d)", "", s)
    s = re.sub(r"(?<=\d)\s*[\.]\s*(?=\d{2}\b)", ".", s)

    # Keep the last amount-like number in case the string contains labels before it.
    matches = re.findall(r"-?\d+(?:\.\d{1,2})?", s)
    if not matches:
        return None

    try:
        return float(matches[-1])
    except ValueError:
        return None


def set_number_format_if_exists(ws, row, headers, header_name, number_format="#,##0.00"):
    col = headers.get(header_name)
    if col:
        ws.cell(row=row, column=col).number_format = number_format


def set_text_format_if_exists(ws, row, headers, header_name):
    col = headers.get(header_name)
    if col:
        ws.cell(row=row, column=col).number_format = "@"

def find_header_row(ws):
    for row in range(1, ws.max_row + 1):
        values = [
            str(ws.cell(row=row, column=col).value).strip()
            for col in range(1, ws.max_column + 1)
            if ws.cell(row=row, column=col).value is not None
        ]

        if "COMPANYCODE" in values and "SUPPLIERINVOICEIDBYINVCGPARTY" in values:
            return row

    raise ValueError(
        "Header row not found. Template must contain COMPANYCODE and SUPPLIERINVOICEIDBYINVCGPARTY."
    )


def build_header_map(ws, header_row):
    headers = {}

    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value

        if value is not None:
            header = str(value).strip()

            # Keep first instance if duplicated.
            if header and header not in headers:
                headers[header] = col

    return headers


def find_next_empty_row(ws, start_row, key_column=1):
    row = start_row

    while ws.cell(row=row, column=key_column).value not in (None, ""):
        row += 1

    return row


def set_if_exists(ws, row, headers, header_name, value=""):
    """Set a worksheet cell when the template contains the target header.

    Some template columns are optional and may intentionally be left blank.
    Defaulting value to an empty string keeps those optional writes from
    failing one-file Excel runs when a caller only needs to reserve the field.
    """
    col = headers.get(header_name)

    if col:
        ws.cell(row=row, column=col).value = value

def set_date_if_exists(ws, row, headers, field_name, value):
    if field_name not in headers or not value:
        return

    cell = ws.cell(row=row, column=headers[field_name])

    if isinstance(value, datetime):
        cell.value = value
    else:
        s = str(value).strip()[:10].replace("-", "/")

        for fmt in ("%m/%d/%Y", "%m/%d/%y"):
            try:
                cell.value = datetime.strptime(s, fmt)
                break
            except ValueError:
                continue
        else:
            cell.value = value

    cell.number_format = "mm/dd/yy"


def write_invoice_row(ws, row, headers, invoice, vendor_config, excel_config, line_number):
    amount = to_number(invoice.get("amount", ""))
    invoice_date = invoice.get("invoice_date") or datetime.today()
    posting_date = datetime.today()
    TaxCenterID = invoice.get("TaxCenterID", "")
    # Always write column A so the record is visible.
    ws.cell(row=row, column=1).value = line_number

    # Also write to any known ID headers if they exist.
    set_if_exists(ws, row, headers, "SUPPLIERINVOICEITEM", line_number)
    set_if_exists(ws, row, headers, "INVOICEID", line_number)
    set_if_exists(ws, row, headers, "ITEMID", line_number)
        
    set_if_exists(ws, row, headers, "SUPPLIERINVOICETRANSACTIONTYPE", excel_config.get("supplier_invoice_transaction_type", "1"), )
    set_if_exists(ws, row, headers, "COMPANYCODE", vendor_config.get("company_code", ""))

    # Invoicing Party / Supplier ID must stay text, so SAP IDs keep leading zeros.
    invoicing_party = str(invoice.get("vendor_id", "")).strip()
    set_if_exists(ws, row, headers, "INVOICINGPARTY", invoicing_party)

    set_if_exists(ws, row, headers, "SUPPLIERINVOICEIDBYINVCGPARTY", as_text(invoice.get("invoice_number", "")))
    
    set_date_if_exists(ws, row, headers, "DOCUMENTDATE", invoice_date)
    set_date_if_exists(ws, row, headers, "POSTINGDATE", posting_date)
    set_if_exists(ws, row, headers, "ACCOUNTINGDOCUMENTTYPE", excel_config.get("accounting_document_type", "NS"))
    #set_if_exists(ws,row,headers,"ACCOUNTINGDOCUMENTHEADERTEXT",f"{invoice.get('vendor_name', '')} Invoice {invoice.get('invoice_number', '')}", )
    set_if_exists(ws, row, headers, "DOCUMENTCURRENCY", excel_config.get("document_currency", "USD"))
    set_if_exists(ws, row, headers, "INVOICEGROSSAMOUNT", amount)

    set_if_exists(ws, row, headers, "GLACCOUNT", (invoice.get("GLAccount") or "51000100"))
    set_if_exists(ws, row, headers, "DEBITCREDITCODE", "S")
    set_if_exists(ws, row, headers, "SUPPLIERINVOICEITEMAMOUNT", amount)
    set_if_exists(ws, row, headers, "TAXCODE", vendor_config.get("tax_code", ""))
    set_if_exists(ws, row, headers, "SUPPLIERINVOICEITEMTEXT", (invoice.get("ItemText") or "Supplies"))
    set_text_format_if_exists(ws, row, headers, "SUPPLIERINVOICEITEMTEXT")
    set_if_exists(ws, row, headers, "TAXJURISDICTION", TaxCenterID)
    set_if_exists(ws, row, headers, "COSTCENTER", vendor_config.get("cost_center", ""))

    # Force SAP upload columns to the correct Excel formats.
    # D = Invoicing Party; K = Gross Invoice Amount; BW = line item amount in this template.
    set_text_format_if_exists(ws, row, headers, "INVOICINGPARTY")
    set_number_format_if_exists(ws, row, headers, "INVOICEGROSSAMOUNT")
    set_number_format_if_exists(ws, row, headers, "SUPPLIERINVOICEITEMAMOUNT")

    for col_letter, number_format in {"D": "@", "K": "#,##0.00", "BW": "#,##0.00"}.items():
        ws[f"{col_letter}{row}"].number_format = number_format

    ws.cell(row=row, column=71).value = vendor_config.get("company_code", "")


def resolve_path(client_root, path_value):
    p = Path(path_value)

    if p.is_absolute() or ":" in str(path_value):
        return p

    return client_root / p
    
