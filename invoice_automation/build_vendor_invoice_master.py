# build_vendor_invoice_master.py
# Separate utility script for North Sky Vendor Invoice (Master).xlsx
# It does NOT modify excel_writer.py, runner.py, or other template workflows.

import os
import re
import shutil
import argparse
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


# -------------------------------------------------------------------
# 1. Try to use your existing repo parser
# -------------------------------------------------------------------
def extract_text_from_pdf(pdf_path):
    """
    Extract text for repo parsers that use parse(text, file_path).
    Uses pdfplumber for text PDFs. If pdfplumber is not available or fails,
    returns empty text and filename fallback still works.
    """
    try:
        import pdfplumber

        parts = []
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text() or ""
                parts.append(page_text)

        return "\n".join(parts).strip()

    except Exception as e:
        print(f"pdfplumber text extraction failed for {pdf_path.name}: {e}")
        return ""


def extract_vendor_invoice_from_filename(pdf_path):
    """
    Filename fallback examples:
      HAT CREEK INV # 3785.pdf -> vendor=HAT CREEK, invoice_number=3785
      D&S SALES 38908.pdf      -> vendor=D&S SALES, invoice_number=38908
    """
    stem = Path(pdf_path).stem
    stem = re.sub(r"\s+", " ", stem).strip()

    patterns = [
        r"^(.+?)\s+INV(?:OICE)?\s*#?\s*([A-Z0-9][A-Z0-9\-\.]*)\b",
        r"^(.+?)\s+#\s*([A-Z0-9][A-Z0-9\-\.]*)\b",
        r"^(.+?)\s+([A-Z0-9][A-Z0-9\-\.]*)$",
    ]

    for pattern in patterns:
        m = re.search(pattern, stem, re.IGNORECASE)
        if m:
            vendor_name = m.group(1).strip()
            invoice_number = m.group(2).strip()
            return vendor_name, invoice_number

    return "", ""



def extract_po_number(text):
    """
    Pattern-match PO number only.
    No bad_values list and no reject logic.
    """

    if not text:
        return ""

    t = text.replace("\r", "\n")
    t = re.sub(r"[ \t]+", " ", t)

    def clean_po(value):
        if value is None:
            return ""

        value = str(value).strip()

        # Stop at next field/header if captured too much.
        value = re.split(
            r"\s{2,}|\n|\b(?:Terms|Freight|Ship|Ship Via|Rep|Project|Job Name|Ordered By|Phone|Email|Delivery Date|Payment Due)\b",
            value,
            maxsplit=1,
            flags=re.IGNORECASE,
        )[0].strip()

        value = value.strip(" :;,#|")
        value = re.sub(
            r"^(No\.?|Number|#)\s*[:\-]?\s*",
            "",
            value,
            flags=re.IGNORECASE,
        ).strip()

        return value.strip(" :;,#|")

    patterns = [
        # Customer PO# 00EQUIP
        # Customer PO # 21621-4000
        # P.O. No: 42767/DUSTIN
        # PO #: JOB#1056255 - PIONEE
        # Order PO #: Bay Area
        r"\b(?:Customer\s+)?P\.?\s*O\.?\s*(?:Number|No\.?|#)?\s*[:#\-]?\s*([A-Z0-9][A-Z0-9/&\-\# ]{0,60})",

        r"\bOrder\s+P\.?\s*O\.?\s*#?\s*[:#\-]?\s*([A-Z0-9][A-Z0-9/&\-\# ]{0,60})",

        # Header then next line:
        # P.O. Number
        # Kirk
        # P.O. No.
        # RESTORATION
        # P.O. No.
        # 21-621-1925
        r"\bP\.?\s*O\.?\s*(?:Number|No\.?|#)\b\s*\n+\s*([A-Z0-9][A-Z0-9/&\-\# ]{0,60})",

        # Table:
        # P.O. NUMBER     TERMS     SHIP     JOB #
        # 2109            NET 30    5/6/2026
        r"\bP\.?\s*O\.?\s*NUMBER\b\s+TERMS\s+SHIP\s+JOB\s*#\s*\n+\s*([A-Z0-9][A-Z0-9/&\-\# ]{0,40})",

        # Table:
        # S.O. No.     P.O. No.     Terms
        # 26497        Fleet        Net 30
        r"\bS\.?\s*O\.?\s*No\.?\s+P\.?\s*O\.?\s*No\.?\s+Terms[\s\S]{0,120}?\n?\s*\d+\s+([A-Z0-9][A-Z0-9/&\-\# ]{0,40})\s+Net\b",
    ]

    for pattern in patterns:
        m = re.search(pattern, t, re.IGNORECASE)
        if m:
            return clean_po(m.group(1))

    return ""


def generic_parse_invoice(pdf_path, text):
    """
    Generic fallback so unknown vendors still produce a row.
    Uses filename for vendor/invoice number and simple regex for date, PO, amount.
    """
    vendor_name, invoice_number = extract_vendor_invoice_from_filename(pdf_path)

    invoice_date = ""
    date_patterns = [
        r"\bInvoice\s+Date\b\s*[:\-]?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
        r"\bDate\s+Invoice\s+No\.?[\s\S]{0,200}?(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
        r"\b(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\b",
    ]

    for pattern in date_patterns:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            invoice_date = m.group(1).strip()
            break

    po_number = extract_po_number(text)

    amount = None
    amount_patterns = [
        r"\bBalance\s+Due\b[\s:.\-]*[$S8]?\s*([0-9,]+\s*[.]\s*\d{2})",
        r"\bAmount\s+Due\b[\s:.\-]*[$S8]?\s*([0-9,]+\s*[.]\s*\d{2})",
        r"\bTotal\s+Due\b[\s:.\-]*[$S8]?\s*([0-9,]+\s*[.]\s*\d{2})",
        r"\bInvoice\s+Total\b[\s:.\-]*[$S8]?\s*([0-9,]+\s*[.]\s*\d{2})",
        r"\bGrand\s+Total\b[\s:.\-]*[$S8]?\s*([0-9,]+\s*[.]\s*\d{2})",
        r"\bTotal\s+Invoice\b[\s\S]{0,500}?([0-9,]+\.\d{2})(?=[\s\r\n]*(?:Internal\s+Ref|TERMS|Email|$))",
    ]

    for pattern in amount_patterns:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            amount = m.group(1).replace(" ", "")
            break

    # Last fallback: use decimal money near bottom, skipping tax/subtotal lines
    if amount is None:
        good_amounts = []
        for line in text.splitlines():
            if re.search(r"\b(tax|sub\s*total|subtotal|freight)\b", line, re.IGNORECASE):
                continue
            good_amounts.extend(re.findall(r"\b([0-9,]+\.\d{2})\b", line))
        if good_amounts:
            amount = good_amounts[-1]

    return {
        "vendor_name": vendor_name,
        "invoice_number": invoice_number,
        "invoice_date": invoice_date,
        "po_number": po_number,
        "amount": amount,
    }


def _call_parser_parse(parser, text, pdf_path):
    """
    Your repo parser signature appears to be parse(text, file_path).
    This helper tries common signatures without changing your parser files.
    """
    try:
        return parser.parse(text, str(pdf_path))
    except TypeError as e1:
        try:
            return parser.parse(text)
        except TypeError:
            try:
                return parser.parse(str(pdf_path))
            except TypeError:
                raise e1


def parse_invoice_with_repo(pdf_path):
    """
    Tries to use existing repo parser logic and returns an invoice dict.
    This script is intentionally separate from your existing SAP Excel writer.

    Fix:
    Your parser error said:
      ValvolineParser.parse() missing 1 required positional argument: 'file_path'

    That means parse() expects text plus file_path, not only pdf_path.
    """
    text = extract_text_from_pdf(pdf_path)

    # Option A: if your repo has detect_parser in app.runner
    try:
        from app.runner import detect_parser

        parser = detect_parser(str(pdf_path))
        invoice = _call_parser_parse(parser, text, pdf_path)
        if invoice:
            return invoice
    except Exception as e:
        print(f"detect_parser failed for {pdf_path.name}: {e}")

    # Option B: if your repo has parse function in runner
    try:
        from app.runner import parse_invoice

        invoice = parse_invoice(str(pdf_path))
        if invoice:
            return invoice
    except Exception as e:
        print(f"parse_invoice failed for {pdf_path.name}: {e}")

    # Option C: default parser fallback if available
    try:
        from app.parsers.valvoline_parser import ValvolineParser

        parser = ValvolineParser()
        invoice = _call_parser_parse(parser, text, pdf_path)
        if invoice:
            return invoice
    except Exception as e:
        print(f"default parser failed for {pdf_path.name}: {e}")

    # Option D: generic fallback from filename + regex text
    invoice = generic_parse_invoice(pdf_path, text)
    if invoice.get("invoice_number") or invoice.get("amount") or invoice.get("invoice_date"):
        print(f"Using generic fallback for {pdf_path.name}")
        return invoice

    return {}


# -------------------------------------------------------------------
# 2. Helpers
# -------------------------------------------------------------------
def clean_amount(value):
    """Return float amount only, even if parser returned '96.10TERMS...'"""
    if value is None:
        return None

    text = str(value).replace(" ", "")
    match = re.search(r"([0-9,]+\.\d{2})", text)
    if not match:
        return None

    try:
        return float(match.group(1).replace(",", ""))
    except ValueError:
        return None


def first_value(invoice, keys):
    for key in keys:
        value = invoice.get(key)
        if value not in [None, ""]:
            return value
    return ""


def build_vendor_display(invoice):
    """Vendor column format: Vendor name - Vendor_id"""
    vendor_name = first_value(invoice, [
        "vendor_name",
        "VendorName",
        "vendor",
        "Vendor",
        "supplier_name",
        "SupplierName",
    ])

    vendor_id = first_value(invoice, [
        "vendor_id",
        "VendorID",
        "supplier",
        "Supplier",
        "supplier_id",
        "SupplierID",
        "company_code",
        "invoicing_party",
        "Invoicing Party",
    ])

    vendor_name = str(vendor_name).strip()
    vendor_id = str(vendor_id).strip()

    if vendor_name and vendor_id:
        return f"{vendor_name} - {vendor_id}"
    if vendor_name:
        return vendor_name
    if vendor_id:
        return vendor_id
    return ""


def normalize_invoice_dict(invoice):
    """Map repo parser output to North Sky Vendor Invoice Master columns."""
    invoice_number = first_value(invoice, [
        "invoice_number",
        "invoice_no",
        "invoice_num",
        "Invoice Number",
        "reference",
        "Reference",
    ])

    invoice_date = first_value(invoice, [
        "invoice_date",
        "Invoice Date",
        "document_date",
        "DOCUMENTDATE",
    ])

    po_number = first_value(invoice, [
        "po_number",
        "po_reference",
        "PO",
        "PO Number",
        "purchase_order",
        "purchase_order_number",
    ])

    amount = first_value(invoice, [
        "amount",
        "total_amount",
        "invoice_amount",
        "Invoice Amount",
        "gross_amount",
        "Gross Invoice Amount in Document Currency",
    ])

    return {
        "Invoice Number": invoice_number,
        "Date Added": datetime.today().strftime("%m/%d/%Y"),
        "Vendor": build_vendor_display(invoice),
        "Invoice Date": invoice_date,
        "Invoice Amount": clean_amount(amount),
        "PO Reference": po_number,
    }


def find_header_row(ws, required_headers):
    """Find row containing at least 3 required headers."""
    max_scan_row = min(ws.max_row, 30)
    for row in range(1, max_scan_row + 1):
        values = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value
            values.append(str(value).strip() if value is not None else "")

        hit_count = sum(1 for header in required_headers if header in values)
        if hit_count >= 3:
            return row

    return None


def prepare_workbook(template_file, output_file):
    """
    Copy North Sky Vendor Invoice (Master).xlsx to output first.
    Original template is never modified.
    """
    required_headers = [
        "Invoice Number",
        "Date Added",
        "Vendor",
        "Invoice Date",
        "Invoice Amount",
        "PO Reference",
    ]

    if not os.path.exists(template_file):
        raise FileNotFoundError(f"Master template not found: {template_file}")

    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    # Overwrite the same output Excel each run by copying the template again.
    shutil.copyfile(template_file, output_file)

    wb = load_workbook(output_file)
    ws = wb.active

    header_row = find_header_row(ws, required_headers)
    if header_row is None:
        raise ValueError(
            "Could not find header row in North Sky Vendor Invoice Master template. "
            "Required headers: Invoice Number, Date Added, Vendor, Invoice Date, Invoice Amount, PO Reference"
        )

    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value:
            headers[str(value).strip()] = col

    missing = [header for header in required_headers if header not in headers]
    if missing:
        raise ValueError(f"Missing required columns in master template: {missing}")

    return wb, ws, header_row, headers


def append_invoice_row(ws, header_row, headers, row_data):
    row = ws.max_row + 1
    if row <= header_row:
        row = header_row + 1

    for header, value in row_data.items():
        col = headers.get(header)
        if col:
            ws.cell(row=row, column=col).value = value

    if "Date Added" in headers:
        ws.cell(row=row, column=headers["Date Added"]).number_format = "mm/dd/yyyy"
    if "Invoice Date" in headers:
        ws.cell(row=row, column=headers["Invoice Date"]).number_format = "mm/dd/yyyy"
    if "Invoice Amount" in headers:
        ws.cell(row=row, column=headers["Invoice Amount"]).number_format = '#,##0.00'

    return row


def get_default_paths(repo_root):
    input_folder = repo_root / "clients" / "northsky_comm" / "downloads"
    template_file = repo_root  / "clients" / "northsky_comm" / "templates" / "North Sky Vendor Invoice (Master).xlsx"

    processed_folder = input_folder / "PROCESSED"
    processed_folder.mkdir(parents=True, exist_ok=True)

    output_file = processed_folder / "North Sky Vendor Invoice (Master)_filled.xlsx"

    return input_folder, template_file, output_file, processed_folder


# -------------------------------------------------------------------
# 3. Main
# -------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Build North Sky Vendor Invoice Master Excel from parsed invoice PDFs."
    )

    parser.add_argument(
        "--input-folder",
        required=False,
        help="Folder containing invoice PDF files. Default: clients/northsky_comm/downloads",
    )
    parser.add_argument(
        "--template",
        required=False,
        help="North Sky Vendor Invoice Master Excel template. Default: templates/North Sky Vendor Invoice (Master).xlsx",
    )
    parser.add_argument(
        "--output",
        required=False,
        help="Output Excel file. Default: downloads/PROCESSED/North Sky Vendor Invoice (Master)_filled_TIMESTAMP.xlsx",
    )
    parser.add_argument(
        "--move-processed",
        action="store_true",
        help="Move successfully parsed PDFs to PROCESSED folder and failed PDFs to BAD folder.",
    )

    args = parser.parse_args()

    repo_root = Path(__file__).resolve().parent
    default_input_folder, default_template_file, default_output_file, processed_folder = get_default_paths(repo_root)

    input_folder = Path(args.input_folder) if args.input_folder else default_input_folder
    template_file = Path(args.template) if args.template else default_template_file
    output_file = Path(args.output) if args.output else default_output_file

    if not input_folder.exists():
        raise FileNotFoundError(f"Input folder not found: {input_folder}")
    if not template_file.exists():
        raise FileNotFoundError(f"Template file not found: {template_file}")

    processed_folder = input_folder / "PROCESSED"
    processed_folder.mkdir(parents=True, exist_ok=True)
    bad_folder = input_folder / "BAD"
    bad_folder.mkdir(parents=True, exist_ok=True)

    pdf_files = sorted(input_folder.glob("*.pdf"))

    print(f"Repo root: {repo_root}")
    print(f"Input folder: {input_folder}")
    print(f"Template file: {template_file}")
    print(f"Output Excel: {output_file}")
    print(f"PDF files found: {len(pdf_files)}")

    wb, ws, header_row, headers = prepare_workbook(str(template_file), str(output_file))

    success_count = 0
    failed_count = 0

    for pdf_path in pdf_files:
        print("=" * 80)
        print(f"Processing: {pdf_path.name}")

        try:
            invoice = parse_invoice_with_repo(pdf_path)
            if not invoice:
                raise ValueError("Parser returned empty invoice data")

            row_data = normalize_invoice_dict(invoice)

            print("Invoice Number:", row_data["Invoice Number"])
            print("Date Added:", row_data["Date Added"])
            print("Vendor:", row_data["Vendor"])
            print("Invoice Date:", row_data["Invoice Date"])
            print("Invoice Amount:", row_data["Invoice Amount"])
            print("PO Reference:", row_data["PO Reference"])

            append_invoice_row(ws, header_row, headers, row_data)
            success_count += 1

            if args.move_processed:
                shutil.move(str(pdf_path), str(processed_folder / pdf_path.name))

        except Exception as e:
            failed_count += 1
            print(f"FAILED {pdf_path.name}: {e}")

            if args.move_processed:
                shutil.move(str(pdf_path), str(bad_folder / pdf_path.name))

    wb.save(output_file)

    print("=" * 80)
    print("Done")
    print(f"Success: {success_count}")
    print(f"Failed: {failed_count}")
    print(f"Output Excel: {output_file}")


if __name__ == "__main__":
    main()
